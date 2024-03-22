import os
import re
from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader

def extract_metadata_from_pdf(filename):
    pdf_file = PdfReader(filename)
    metadata = {
        "Autor": pdf_file.metadata.author,
        "Titulo": pdf_file.metadata.title,
        "Fecha de Creación": pdf_file.metadata.get('/CreationDate', None),
        "Modificado": pdf_file.metadata.get('/ModDate', None)
    }
    return metadata

def extract_metadata_from_docx(filename):
    document = Document(filename)
    core_properties = document.core_properties
    metadata = {
        "Autor": core_properties.author,
        "Titulo": core_properties.title,
        "Fecha de Creación": core_properties.created,
        "Modificado": core_properties.modified
    }
    return metadata

def extract_metadata_from_xlsx(filename):
    wb = load_workbook(filename)
    author = None
    #verificamos la información para revisar si existe algún autor
    if wb.sheetnames:
        sheet = wb[wb.sheetnames[0]]
        if sheet.dimensions:
            author = sheet.cell(row=1, column=1).value
    metadata = {
        "Autor": author,
        "Titulo": wb.properties.title,
        "Fecha de Creación": wb.properties.created,
        "Modificado": wb.properties.modified
    }
    return metadata

def process_files(directory_path):
    for file_name in os.listdir(directory_path):
        file_path = os.path.join(directory_path, file_name)
        if os.path.isfile(file_path):
            file_extension = re.findall(r"\.(pdf|docx|xlsx)$", file_name)
            if file_extension:
                file_extension = file_extension[0]
                if file_extension == "pdf":
                    metadata = extract_metadata_from_pdf(file_path)
                elif file_extension == "docx":
                    metadata = extract_metadata_from_docx(file_path)
                elif file_extension == "xlsx":
                    metadata = extract_metadata_from_xlsx(file_path)
                print(f"Metadatos del archivo {file_name}:")
                for k, v in metadata.items():
                    print(f"  {k}: {v}")
                print()

# Ruta al directorio
directory_path = "/home/emmanuel/Desktop/CyberSecurity/Tarea2"

# Procesar los archivos en el directorio
process_files(directory_path)
